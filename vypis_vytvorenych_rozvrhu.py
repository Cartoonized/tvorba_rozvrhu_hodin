import openpyxl
import copy
# from typing import Union


# rozvrh_type = list[list[list[list[Union[list[int], list[[]]]]]]]
rozvrh_type = list


def Vypis_Rozvrh(rozvrhy: list[rozvrh_type], seznam_oboru: list[str], seznam_predmetu: list[str], seznam_mistnosti: list[str], seznam_vyucujicich: list[str]) -> None:
    """
    Výpis rozvrhu do console.
    :param rozvrhy: seznam vytvořených rozvrhů
    :param seznam_oboru: seznam oborů
    :param seznam_predmetu: seznam předmětů
    :param seznam_mistnosti: seznam místností
    :param seznam_vyucujicich: seznam vyučujících
    :return: None
    """
    print(rozvrhy)
    rozvrhy = copy.deepcopy(rozvrhy)
    for rozvrh in rozvrhy:
        print("")
        for obor in range(len(seznam_oboru)):
            for den in range(5):
                for cas in range(12):
                    if rozvrh[obor][den][cas]:
                        rozvrh[obor][den][cas][0] = seznam_vyucujicich[rozvrh[obor][den][cas][0]]
                        rozvrh[obor][den][cas][1] = seznam_predmetu[rozvrh[obor][den][cas][1]]
                        rozvrh[obor][den][cas][2] = seznam_mistnosti[rozvrh[obor][den][cas][2]]
        for obor in range(len(seznam_oboru)):
            print(f"{seznam_oboru[obor]:}")
            for den in range(5):
                print(f"{rozvrh[obor][den]}")
            print("")

"""
def Vystup(cesta: str, rozvrhy: list[rozvrh_type], seznam_oboru: list[str], seznam_predmetu: list[str], seznam_mistnosti: list[str], seznam_vyucujicich: list[str]) -> None:

    #výpis rozvrhů do XLSX souboru
    #:param cesta: cesta k souboru XLXS
    #:param rozvrhy: seznam vytvořených rozvrhů
    #:param seznam_oboru: seznam oborů
    #:param seznam_predmetu: seznam předmětů
    #:param seznam_mistnosti: seznam místností
    #:param seznam_vyucujicich: seznam vyučujících
    #:return: None

    try:
        book = openpyxl.load_workbook(cesta)
    except FileNotFoundError:
        book = openpyxl.Workbook()
        book.remove(book.active)

    barvy = ['#FF9999']
    dvt = ["Po", "Út", "St", "Čt", "Pá"]
    index = 0

    for rozvrh in rozvrhy:
        for obor in range(len(seznam_oboru)):
            if seznam_oboru[obor] in book.sheetnames:
                sheet = book[seznam_oboru[obor]]
            else:
                sheet = book.create_sheet(seznam_oboru[obor])

            for den in range(5):
                index_barvy = 0

                cell = sheet.cell(row=1 + (index * 6) + (index), column=1)
                if index == 0:
                    cell = sheet.cell(row=1 + (index * 6), column=1)
                cell.value = index + 1

                cell = sheet.cell(row=den + 2 + (index*6) + (index), column=1)
                if index == 0:
                    cell = sheet.cell(row=den + 2 + (index * 6), column=1)
                cell.value = dvt[den]

                for cas in range(len(rozvrh[obor][den])):
                    cell = sheet.cell(row=1 + (index*6) + (index), column=cas + 2)
                    if index == 0:
                        cell = sheet.cell(row=1 + (index*6), column=cas + 2)
                    cell.value = f"{cas+8}-{cas+9}"

                    if rozvrh[obor][den][cas]:
                        vyucujici = seznam_vyucujicich[rozvrh[obor][den][cas][0]]
                        mistnost = seznam_mistnosti[rozvrh[obor][den][cas][2]]
                        predmet = seznam_predmetu[rozvrh[obor][den][cas][1]]
                        hodnoty = f"{predmet, mistnost, vyucujici}"

                        # cell = sheet.cell(row=den + (index * 5) + 2, column=cas + 2)
                        cell = sheet.cell(row=den + 2 + index*6 + (index), column=cas + 2)
                        if index == 0:
                            cell = sheet.cell(row=den + 2, column=cas + 2)
                        cell.value = hodnoty
                        cell.fill = openpyxl.styles.PatternFill(start_color=barvy[index_barvy][1:],
                                                                end_color=barvy[index_barvy][1:],
                                                                fill_type="solid")
            sirka_sloupce = 30
            for col_num in range(2, 14):  # Sloupce 2 až 13, protože cas_od začíná na sloupci 1
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = sirka_sloupce
        # book.save(cesta)
        index += 1
    # Uložení zpět do souboru
    book.save(cesta)
    book.close()
"""

