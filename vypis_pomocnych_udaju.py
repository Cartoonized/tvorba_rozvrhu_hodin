import openpyxl


def Vypis_PomUdajeDoExcelu(cesta: str, pom_seznam, seznam_predmetu):
    """
    Vypisuje pomocná data do EXCELu.
    :param cesta: cesta k souboru
    :param pom_seznam: seznam pomocných dat
    :param seznam_predmetu: seznam předmětů pro převod indexů zpět na názvy
    :return: None
    """
    listy = ["Vlivy", "Predmety", "Hodnoty", "Nez.Predmety"]
    if pom_seznam:
        """
        try:
            book = openpyxl.load_workbook(cesta)
        except FileNotFoundError:
            book = openpyxl.Workbook()
            book.remove(book.active)
        """
        book = openpyxl.Workbook()
        book.remove(book.active)

        for list in range(4):
            sheet = book.create_sheet(listy[list])
            for radek in range(len(pom_seznam)):
                if list != 2:
                    if list != 3:
                        for hodnota in range(len(pom_seznam[radek][list])):
                            cell = sheet.cell(row=radek+1, column=hodnota+1)
                            cell.value = str(pom_seznam[radek][list][hodnota])
                    else:
                        for hodnota in range(len(pom_seznam[radek][list])):
                            cell = sheet.cell(row=radek+1, column=hodnota+1)
                            cell.value = str([seznam_predmetu[pom_seznam[radek][list][hodnota][0]], str(pom_seznam[radek][list][hodnota][1])])  # [X, Y] (předmět, počet_přednášek)
                else:
                    for predmet in range(len(seznam_predmetu)):
                        cell = sheet.cell(row=1, column=predmet+1)
                        cell.value = str(seznam_predmetu[predmet])
                    for hodnota in range(len(pom_seznam[radek][list])):
                        cell = sheet.cell(row=radek+2, column=hodnota+1)
                        cell.value = str(pom_seznam[radek][list][hodnota])
        book.save(cesta)
        book.close()